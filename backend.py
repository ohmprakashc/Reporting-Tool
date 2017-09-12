# -*- coding: utf-8 -*-
"""
Created on Sun Sep  3 21:15:20 2017

@author: Toshiba
"""

import openpyxl as xl
import difflib
from os import listdir
from os.path import isfile, join
#==============================================================================
# create header data
header=('Resource','DOJ','Vendor','BU','Manager','Tech','Skill',
        'Area of Service','Location','Experience','billable days','Old rate card',
        'New rate card','billing')

print ("-------------------------------------------------------")
print ("WARNING: Make sure there is no file by name header.xlsx")
print ("-------------------------------------------------------")
input("Click any key to continue")
# Create excel sheet
try:
    wb=xl.Workbook()
    wb.save('header.xlsx')
    print ("Creating excel sheet success")
except:
    print("Creating excel sheet failed")

# Open work sheet
row=1
try:
    sheet=wb.get_sheet_by_name('Sheet')
    print ("Opening excel sheet success")
except:
    print("Opening excel sheet failed")

# Write data to work sheet
try:
    for i in range(1,len(header)):
        sheet.cell(row=1,column=i).value=header[i-1]
    wb.save('header.xlsx')
    wb.close()
    print ("Writting header data to excel file success")
except:
    wb.save('header.xlsx')
    print("Writting header data to excel file failed")
#==============================================================================

#==============================================================================
# Filter xlsx files from directory provided by user

try:
    mypath=input ('Enter path of directory where files has to be searched\n')
except:
    print ("Unable to get input from keyboard")
#listfiles= os.listdir()

try:
    list_all_files=[f for f in listdir(mypath) if isfile(join(mypath, f))]
    #listfiles= (difflib.get_close_matches('.xlsx',listfiles,n=len(listfiles),cutoff=0.4))
    xlsxFileFilter= [f for f in list_all_files if ".xlsx" in f]
    xlsxFileFilter.remove('header.xlsx')
    
    if len(xlsxFileFilter) ==0:
        print ("No XLSX files found. Check the path again")
    else:
        print ("List of XLSX files that will be consolidated", xlsxFileFilter)
except:
    print ("Unable to read files in given path.")
#==============================================================================

#==============================================================================
# Get header from first file & compare with header file header
# If both headers match, copy paste the column data
# Alternately try to use data frames or xlwings

# Gets first row of given excel sheet
def iter_row(sheet):
    for row in sheet.iter_rows():
        yield [cell.value for cell in row]

for fileName in xlsxFileFilter:
    
    try:
        headerFile=xl.load_workbook('header.xlsx')
        headerSheet=headerFile.get_sheet_by_name('Sheet')
        print ("Opened header.xlsx file successfully")
        headerSheetMaxColumn=headerSheet.max_column
        headerSheetMaxRow=headerSheet.max_row
    except:
        print ("Unable to open header.xlsx file")
        break
    
    try:
        supplierFile=xl.load_workbook(fileName)
        print ("Opened file", fileName, "successfully for processing")
    except:
        print ("Unable to open file for processing",fileName)    
        supplierFile.close()
        break
    
    try:    
        sheetInSupplierFile=supplierFile.get_sheet_by_name('Sheet1')
        supplierFileMaxRow=sheetInSupplierFile.max_row
        supplierFileMaxColumn=sheetInSupplierFile.max_column
        print (supplierFileMaxRow,supplierFileMaxColumn)
        print("Opened sheet in supplier file",fileName)
    except:
        print ("Unable to open Sheet 1 in", fileName, "file for processing")
        supplierFile.close()
        break
    
    for headerCounter in range(1,headerSheetMaxColumn):
        
        print ("Header=",headerSheet.cell(row=1,column=headerCounter).value)
        SupplierHeader=(list(iter_row(sheetInSupplierFile))[0])
        print (SupplierHeader)
        
        try:
            headerValue=headerSheet.cell(row=1,column=headerCounter).value
            bestmatch=difflib.get_close_matches(headerValue,SupplierHeader,1)[0]
            print ("Best match for header=",bestmatch)
            #matchingColInSupplierFile=SupplierHeader.index(bestmatch)
            print ("Best match is in column=",SupplierHeader.index(bestmatch))
        except:
            print ("No match found")
            break
headerFile.save('header.xlsx')
headerFile.close()
print ("Completed!")